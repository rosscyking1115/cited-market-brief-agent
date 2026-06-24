import base64
import binascii
import csv
import io
import re
from datetime import date, datetime
from zoneinfo import ZoneInfo

from openpyxl import load_workbook

from app.connectors.twse import TwseClient, fetch_listed_industry_map
from app.fund_attribution.schemas import (
    AttributionRow,
    AutomationPolicyItem,
    BenchmarkReturnOut,
    BenchmarkReturnRequest,
    FundAttributionOut,
    FundAttributionPlanOut,
    FundAttributionRequest,
    FundConfig,
    FundReturnOut,
    FundReturnRequest,
    HoldingInput,
    HoldingReturnFillOut,
    HoldingReturnFillRequest,
    HoldingsFileParseRequest,
    HoldingsParseOut,
    HoldingsParseRequest,
    SectorAttributionOut,
    SectorAttributionRow,
    SectorConfig,
)
from app.fund_attribution.store import load_config, load_sector_config, save_result

TAIPEI_TZ = ZoneInfo("Asia/Taipei")

DISCLAIMER = "本頁提供績效歸因與教育資訊，不構成個人化投資建議或買賣建議。"

COLUMN_ALIASES = {
    "symbol": {
        "symbol",
        "ticker",
        "ticker_symbol",
        "stock_code",
        "code",
        "sedol",
        "isin",
        "ric",
        "代號",
        "股票代碼",
        "股票代號",
        "證券代號",
        "標的代號",
    },
    "name": {
        "name",
        "security",
        "security_name",
        "holding",
        "holdings",
        "company",
        "issuer",
        "description",
        "名稱",
        "股票名稱",
        "證券名稱",
        "持股名稱",
        "基金資產股票",
    },
    "weight_pct": {
        "weight",
        "weight_pct",
        "weight_percent",
        "weighting",
        "percent",
        "pct",
        "of_nav",
        "nav_percent",
        "market_value_percent",
        "權重",
        "比重",
        "持股比重",
        "權重_percent",
        "占比",
        "淨資產百分比",
    },
    "return_pct": {
        "return",
        "return_pct",
        "return_percent",
        "daily_return",
        "daily_return_pct",
        "漲跌幅",
        "報酬率",
        "日報酬",
        "當日報酬",
    },
    "sector": {
        "sector",
        "industry",
        "gics_sector",
        "產業",
        "產業別",
        "類股",
        "類別",
        "產業類別",
    },
}


def automation_policy() -> list[AutomationPolicyItem]:
    return [
        AutomationPolicyItem(
            label="JPM ETF holdings file",
            status="manual_first",
            note=(
                "先支援使用者上傳官網下載的持股檔；自動下載前需確認 JPM 頁面條款與機器存取方式。"
            ),
        ),
        AutomationPolicyItem(
            label="TWSE after-close prices",
            status="allowed",
            note="台股每日收盤後可用官方公開資料抓取個股收盤價；需保留來源、時間與快取紀錄。",
        ),
        AutomationPolicyItem(
            label="TAIEX benchmark close",
            status="needs_review",
            note="可用官方或授權來源做每日比較；公開產品前需確認指數顯示/再發布條款。",
        ),
        AutomationPolicyItem(
            label="Intraday attribution",
            status="blocked",
            note="盤中即時歸因通常需要正式行情授權；第一版只做每日收盤後分析。",
        ),
    ]


def attribution_plan() -> FundAttributionPlanOut:
    return FundAttributionPlanOut(
        title="ETF / Fund vs Benchmark Attribution",
        target_use_case="每天收盤後回答：基金和大盤表現差在哪裡，哪些持股造成差異。",
        first_region="Taiwan",
        daily_trigger="台股收盤後，等 TWSE/JPM 資料更新再跑。",
        required_inputs=[
            "ETF 持股、權重與持股日期",
            "ETF 當日報酬率或淨值/收盤價",
            "台灣加權指數當日報酬率",
            "每個持股的當日報酬率",
        ],
        first_supported_workflow=[
            "先讓使用者上傳 JPM 官網下載的持股檔",
            "系統抓 TWSE 收盤價並計算每檔持股報酬",
            "用權重 x 報酬計算貢獻",
            "輸出贏/輸大盤原因、最大貢獻、最大拖累、無法解釋殘差",
        ],
        automation_policy=automation_policy(),
        disclaimer=DISCLAIMER,
    )


def parse_holdings_text(request: HoldingsParseRequest) -> HoldingsParseOut:
    rows, detected_columns = _read_tabular_text(request.text)
    return _holdings_out_from_rows(
        rows=rows,
        detected_columns=detected_columns,
        source_name=request.source_name,
        source_note="manual holdings parse",
    )


def parse_holdings_workbook(request: HoldingsFileParseRequest) -> HoldingsParseOut:
    source_name = request.source_name or request.filename
    try:
        content = base64.b64decode(request.content_base64, validate=True)
    except (binascii.Error, ValueError):
        return HoldingsParseOut(
            source_name=source_name,
            parsed_count=0,
            skipped_rows=0,
            detected_columns=[],
            holdings=[],
            warnings=["檔案內容不是有效的 base64；請重新選擇 Excel 檔。"],
            source_notes=[f"source: {source_name}", "JPM holdings workbook parse failed"],
        )

    try:
        rows, detected_columns, metadata, warnings = _read_workbook_holdings(content)
    except Exception as exc:  # noqa: BLE001 - bad uploads should become user-facing warnings.
        return HoldingsParseOut(
            source_name=source_name,
            parsed_count=0,
            skipped_rows=0,
            detected_columns=[],
            holdings=[],
            warnings=[f"Excel 檔解析失敗：{exc}"],
            source_notes=[f"source: {source_name}", "JPM holdings workbook parse failed"],
        )

    parsed = _holdings_out_from_rows(
        rows=rows,
        detected_columns=detected_columns,
        source_name=source_name,
        source_note="JPM holdings workbook parse",
        as_of=metadata.get("as_of"),
        fund_name=metadata.get("fund_name"),
    )
    return parsed.model_copy(update={"warnings": [*warnings, *parsed.warnings]})


def _holdings_out_from_rows(
    *,
    rows: list[dict[str, str]],
    detected_columns: list[str],
    source_name: str,
    source_note: str,
    as_of: str | None = None,
    fund_name: str | None = None,
) -> HoldingsParseOut:
    warnings: list[str] = []
    skipped = 0
    holdings: list[HoldingInput] = []

    if not rows:
        return HoldingsParseOut(
            source_name=source_name,
            as_of=as_of,
            fund_name=fund_name,
            parsed_count=0,
            skipped_rows=0,
            detected_columns=detected_columns,
            holdings=[],
            warnings=["找不到可解析的持股表格。請貼上 CSV/TSV 或從試算表複製完整表格。"],
            source_notes=[f"source: {source_name}", source_note],
        )

    for index, row in enumerate(rows, start=1):
        symbol = _clean_text(row.get("symbol"))
        name = _clean_text(row.get("name"))
        weight = _parse_percent(row.get("weight_pct"))
        return_pct = _parse_percent(row.get("return_pct"))
        sector = _clean_text(row.get("sector")) or None

        if not symbol and name:
            symbol = _infer_symbol(name)
        if not name and symbol:
            name = symbol
        if not name or weight is None:
            skipped += 1
            continue

        holdings.append(
            HoldingInput(
                symbol=symbol or f"ROW-{index}",
                name=name,
                weight_pct=weight,
                return_pct=return_pct,
                sector=canonical_sector(sector) if sector else None,
            )
        )

    if not holdings:
        warnings.append("沒有成功解析任何持股；請確認欄位包含名稱與權重。")
    if skipped:
        warnings.append(f"已略過 {skipped} 列缺少名稱或權重的資料。")

    return HoldingsParseOut(
        source_name=source_name,
        as_of=as_of,
        fund_name=fund_name,
        parsed_count=len(holdings),
        skipped_rows=skipped,
        detected_columns=detected_columns,
        holdings=holdings,
        warnings=warnings,
        source_notes=[f"source: {source_name}", source_note],
    )


def fill_holding_returns_from_twse(request: HoldingReturnFillRequest) -> HoldingReturnFillOut:
    as_of = _parse_iso_date(request.as_of)
    if as_of is None:
        return HoldingReturnFillOut(
            as_of=request.as_of,
            filled_count=0,
            missing_symbols=[holding.symbol for holding in request.holdings],
            holdings=request.holdings,
            warnings=["日期格式需要是 YYYY-MM-DD。"],
            source_notes=["TWSE after-close return fill skipped"],
        )

    client = TwseClient()
    holdings: list[HoldingInput] = []
    missing: list[str] = []
    warnings: list[str] = []
    filled = 0
    try:
        for holding in request.holdings:
            symbol = _normalize_twse_symbol(holding.symbol)
            if not symbol:
                holdings.append(holding)
                missing.append(holding.symbol)
                continue
            try:
                daily_return = client.stock_daily_return(symbol=symbol, as_of=as_of)
            except Exception as exc:  # noqa: BLE001 - source failures should not break manual flow.
                warnings.append(f"{symbol} TWSE 取價失敗：{exc}")
                holdings.append(holding)
                missing.append(holding.symbol)
                continue
            if daily_return is None:
                holdings.append(holding)
                missing.append(holding.symbol)
                continue
            holdings.append(holding.model_copy(update={"return_pct": daily_return.return_pct}))
            filled += 1
    finally:
        client.close()

    if missing:
        warnings.append(f"有 {len(missing)} 檔沒有補到 TWSE 漲跌幅，仍可手動填入。")

    return HoldingReturnFillOut(
        as_of=as_of.isoformat(),
        filled_count=filled,
        missing_symbols=missing,
        holdings=holdings,
        warnings=warnings,
        source_notes=[
            "source: TWSE afterTrading STOCK_DAY",
            "returns calculated from latest available close and previous close",
        ],
    )


def benchmark_return_from_twse(request: BenchmarkReturnRequest) -> BenchmarkReturnOut:
    as_of = _parse_iso_date(request.as_of)
    if as_of is None:
        return BenchmarkReturnOut(
            as_of=request.as_of,
            benchmark=request.benchmark,
            name="台灣加權指數",
            return_pct=None,
            close=None,
            previous_close=None,
            warnings=["日期格式需要是 YYYY-MM-DD。"],
            source_notes=["TWSE benchmark return skipped"],
        )

    client = TwseClient()
    warnings: list[str] = []
    try:
        try:
            value = client.taiex_return(as_of=as_of)
        except Exception as exc:  # noqa: BLE001 - keep manual benchmark fallback available.
            value = None
            warnings.append(f"TAIEX TWSE 取價失敗：{exc}")
    finally:
        client.close()

    if value is None:
        warnings.append("沒有補到台灣加權指數漲跌幅，仍可手動填入。")
        return BenchmarkReturnOut(
            as_of=as_of.isoformat(),
            benchmark=request.benchmark,
            name="台灣加權指數",
            return_pct=None,
            close=None,
            previous_close=None,
            warnings=warnings,
            source_notes=["source: TWSE afterTrading MI_INDEX"],
        )

    return BenchmarkReturnOut(
        as_of=value.trade_date,
        benchmark=value.symbol,
        name=value.name,
        return_pct=value.return_pct,
        close=value.close,
        previous_close=value.previous_close,
        warnings=warnings,
        source_notes=[
            "source: TWSE afterTrading MI_INDEX",
            "benchmark return from TWSE reported index percentage change",
        ],
    )


def fund_return_from_twse(request: FundReturnRequest) -> FundReturnOut:
    """Fetch the ETF/fund's own daily return from TWSE by its listed stock code.

    Taiwan active ETFs carry a letter-suffixed code (e.g. 00982A), so the symbol
    normaliser keeps a trailing letter. Manual entry stays available on failure.
    """
    as_of = _parse_iso_date(request.as_of)
    symbol = _normalize_twse_fund_symbol(request.symbol)
    if as_of is None or not symbol:
        return FundReturnOut(
            as_of=request.as_of,
            symbol=symbol or request.symbol,
            name=request.symbol,
            return_pct=None,
            close=None,
            previous_close=None,
            warnings=["請輸入有效日期 (YYYY-MM-DD) 與基金/ETF 在台股的代號。"],
            source_notes=["TWSE fund return skipped"],
        )

    client = TwseClient()
    warnings: list[str] = []
    try:
        try:
            value = client.stock_daily_return(symbol=symbol, as_of=as_of)
        except Exception as exc:  # noqa: BLE001 - keep manual fund-return fallback available.
            value = None
            warnings.append(f"{symbol} TWSE 取價失敗：{exc}")
    finally:
        client.close()

    if value is None:
        warnings.append("沒有補到基金/ETF 當日漲跌幅，請確認代號或手動填入。")
        return FundReturnOut(
            as_of=as_of.isoformat(),
            symbol=symbol,
            name=symbol,
            return_pct=None,
            close=None,
            previous_close=None,
            warnings=warnings,
            source_notes=["source: TWSE afterTrading STOCK_DAY"],
        )

    return FundReturnOut(
        as_of=value.trade_date,
        symbol=value.symbol,
        name=value.symbol,
        return_pct=value.return_pct,
        close=value.close,
        previous_close=value.previous_close,
        warnings=warnings,
        source_notes=[
            "source: TWSE afterTrading STOCK_DAY",
            "fund/ETF return from latest available close vs previous close",
        ],
    )


def analyze_fund_attribution(request: FundAttributionRequest) -> FundAttributionOut:
    rows = [_row_from_holding(holding) for holding in request.holdings]
    explained_return = sum(row.contribution_pct or 0 for row in rows)
    active_return = request.fund_return_pct - request.benchmark_return_pct
    residual = request.fund_return_pct - explained_return

    contributors = sorted(
        [row for row in rows if (row.contribution_pct or 0) > 0],
        key=lambda row: row.contribution_pct or 0,
        reverse=True,
    )[:5]
    drags = sorted(
        [row for row in rows if (row.contribution_pct or 0) < 0],
        key=lambda row: row.contribution_pct or 0,
    )[:5]
    missing = [row for row in rows if row.direction == "missing"]

    return FundAttributionOut(
        fund_name=request.fund_name,
        benchmark_name=request.benchmark_name,
        as_of=request.as_of,
        fund_return_pct=round(request.fund_return_pct, 4),
        benchmark_return_pct=round(request.benchmark_return_pct, 4),
        active_return_pct=round(active_return, 4),
        explained_return_pct=round(explained_return, 4),
        residual_pct=round(residual, 4),
        holdings_count=len(request.holdings),
        contributors=contributors,
        drags=drags,
        missing_returns=missing,
        source_notes=request.source_notes,
        automation_policy=automation_policy(),
        summary_zh_hant=_summary_zh(
            fund_name=request.fund_name,
            benchmark_name=request.benchmark_name,
            fund_return=request.fund_return_pct,
            benchmark_return=request.benchmark_return_pct,
            active_return=active_return,
            contributors=contributors,
            drags=drags,
        ),
        disclaimer=DISCLAIMER,
    )


def refresh_latest_attribution(as_of: date | None = None) -> FundAttributionOut | None:
    """Recompute today's attribution from the saved config, using TWSE after-close
    data, and persist it so the page can show it pre-computed. Returns None when no
    fund has been configured yet. Designed to run unattended on an evening cron."""
    config = load_config()
    if config is None or not config.holdings:
        return None

    day = as_of or datetime.now(TAIPEI_TZ).date()
    iso = day.isoformat()
    notes: list[str] = list(config.source_notes)

    fill = fill_holding_returns_from_twse(
        HoldingReturnFillRequest(as_of=iso, holdings=config.holdings)
    )
    holdings = fill.holdings
    notes.extend(fill.source_notes)

    benchmark = benchmark_return_from_twse(BenchmarkReturnRequest(as_of=iso, benchmark="TAIEX"))
    benchmark_return = benchmark.return_pct
    notes.extend(benchmark.source_notes)

    fund_return: float | None = None
    if config.fund_symbol:
        fund = fund_return_from_twse(FundReturnRequest(as_of=iso, symbol=config.fund_symbol))
        fund_return = fund.return_pct
        notes.extend(fund.source_notes)

    explained = sum(
        holding.weight_pct * holding.return_pct / 100
        for holding in holdings
        if holding.return_pct is not None
    )
    if fund_return is None:
        fund_return = round(explained, 4)
        notes.append("基金當日報酬以持股貢獻估算（未取得 ETF 收盤價）。")
    if benchmark_return is None:
        benchmark_return = 0.0
        notes.append("未取得台灣加權指數當日報酬，暫以 0 計算。")

    result = analyze_fund_attribution(
        FundAttributionRequest(
            fund_name=config.fund_name,
            benchmark_name=config.benchmark_name,
            as_of=iso,
            fund_return_pct=fund_return,
            benchmark_return_pct=benchmark_return,
            holdings=holdings,
            source_notes=notes,
        )
    )
    save_result(result)
    return result


def save_fund_config(config: FundConfig) -> FundConfig:
    from app.fund_attribution.store import save_config

    save_config(config)
    return config


# --- Sector (產業) attribution -------------------------------------------------

_SECTOR_SUFFIXES = ("類股價指數", "類報酬指數", "類指數", "類股指數", "類股", "類", "業")


def canonical_sector(name: str) -> str:
    """Normalise a sector label so the same sector matches across sources: the
    holdings file (e.g. 半導體業), the TWSE sector index (半導體類指數), and the
    stored TAIEX weights (半導體) all collapse to one key."""
    cleaned = (name or "").strip().replace(" ", "")
    for suffix in _SECTOR_SUFFIXES:
        if len(cleaned) > len(suffix) and cleaned.endswith(suffix):
            cleaned = cleaned[: -len(suffix)]
            break
    return cleaned or name.strip()


def _holding_sector(holding: HoldingInput, sector_map: dict[str, str]) -> str | None:
    if holding.sector and holding.sector.strip():
        return canonical_sector(holding.sector)
    mapped = sector_map.get(holding.symbol)
    return canonical_sector(mapped) if mapped and mapped.strip() else None


def aggregate_etf_sectors(
    holdings: list[HoldingInput], sector_map: dict[str, str]
) -> tuple[dict[str, float], float]:
    weights: dict[str, float] = {}
    unmapped = 0.0
    for holding in holdings:
        sector = _holding_sector(holding, sector_map)
        if sector is None:
            unmapped += holding.weight_pct
        else:
            weights[sector] = weights.get(sector, 0.0) + holding.weight_pct
    return weights, round(unmapped, 4)


def build_sector_attribution(
    *,
    holdings: list[HoldingInput],
    sector_returns: dict[str, float],
    config: SectorConfig,
    fund_name: str,
    benchmark_name: str,
    as_of: str,
    source_notes: list[str] | None = None,
) -> SectorAttributionOut:
    etf_weights, unmapped = aggregate_etf_sectors(holdings, config.sector_map)
    taiex = {canonical_sector(w.sector): w.weight_pct for w in config.taiex_weights}
    returns = {canonical_sector(name): value for name, value in sector_returns.items()}
    has_benchmark = bool(taiex)

    sectors = sorted(
        set(etf_weights) | set(taiex),
        key=lambda s: etf_weights.get(s, 0.0),
        reverse=True,
    )
    rows: list[SectorAttributionRow] = []
    alloc_total: float | None = 0.0 if has_benchmark else None
    for sector in sectors:
        ew = round(etf_weights.get(sector, 0.0), 4)
        bw = taiex.get(sector)
        ret = returns.get(sector)
        diff = round(ew - bw, 4) if bw is not None else None
        etf_contrib = round(ew * ret / 100, 4) if ret is not None else None
        alloc = round((ew - bw) * ret / 100, 4) if (bw is not None and ret is not None) else None
        if alloc is not None and alloc_total is not None:
            alloc_total += alloc
        rows.append(
            SectorAttributionRow(
                sector=sector,
                etf_weight_pct=ew,
                benchmark_weight_pct=bw,
                weight_diff_pct=diff,
                sector_return_pct=ret,
                etf_contribution_pct=etf_contrib,
                allocation_effect_pct=alloc,
            )
        )
    if alloc_total is not None:
        alloc_total = round(alloc_total, 4)

    return SectorAttributionOut(
        as_of=as_of,
        fund_name=fund_name,
        benchmark_name=benchmark_name,
        has_benchmark=has_benchmark,
        rows=rows,
        allocation_total_pct=alloc_total,
        unmapped_weight_pct=unmapped,
        summary_zh_hant=_sector_summary_zh(
            rows=rows,
            benchmark_name=benchmark_name,
            has_benchmark=has_benchmark,
            alloc_total=alloc_total,
        ),
        source_notes=source_notes or [],
        disclaimer=DISCLAIMER,
    )


def compute_sector_attribution(as_of: date | None = None) -> SectorAttributionOut:
    """Live sector attribution from the saved fund holdings + sector config and
    today's TWSE sector-index returns. Returns an empty (but valid) result when no
    fund has been configured yet."""
    day = as_of or datetime.now(TAIPEI_TZ).date()
    iso = day.isoformat()
    fund = load_config()
    if fund is None or not fund.holdings:
        return SectorAttributionOut(
            as_of=iso,
            fund_name="",
            benchmark_name="台灣加權指數",
            has_benchmark=False,
            rows=[],
            allocation_total_pct=None,
            unmapped_weight_pct=0.0,
            summary_zh_hant="尚未設定基金；請先在持股工具按「設為每日自動更新」。",
            source_notes=[],
            disclaimer=DISCLAIMER,
        )

    notes: list[str] = []
    sector_returns: dict[str, float] = {}
    client = TwseClient()
    try:
        sector_returns = client.sector_returns(as_of=day)
    except Exception as exc:  # noqa: BLE001 - keep the breakdown usable without returns.
        notes.append(f"TWSE 產業指數取得失敗：{exc}")
    finally:
        client.close()
    if sector_returns:
        notes.append("source: TWSE afterTrading MI_INDEX 類股指數")

    # Auto-classify holdings via TWSE's public industry data; the stored map (and
    # any 產業別 column already on a holding) takes precedence.
    stored = load_sector_config()
    auto_map = fetch_listed_industry_map()
    if auto_map:
        notes.append("產業分類：TWSE 上市公司產業別")
    config = SectorConfig(
        taiex_weights=stored.taiex_weights,
        sector_map={**auto_map, **stored.sector_map},
    )

    return build_sector_attribution(
        holdings=fund.holdings,
        sector_returns=sector_returns,
        config=config,
        fund_name=fund.fund_name,
        benchmark_name=fund.benchmark_name,
        as_of=iso,
        source_notes=notes,
    )


def _sector_summary_zh(
    *,
    rows: list[SectorAttributionRow],
    benchmark_name: str,
    has_benchmark: bool,
    alloc_total: float | None,
) -> str:
    if not has_benchmark or alloc_total is None:
        scored = [r for r in rows if r.etf_contribution_pct is not None]
        if not scored:
            return "已彙整基金產業配置；補上各產業當日漲跌幅後即可看到每個產業的貢獻。"
        top = max(scored, key=lambda r: r.etf_contribution_pct or 0)
        bottom = min(scored, key=lambda r: r.etf_contribution_pct or 0)
        return (
            f"本日基金產業貢獻：{top.sector} 最為正貢獻、{bottom.sector} 最為拖累。"
            f"設定加權指數產業權重後，可進一步比較配置差異。"
        )
    effects = [r for r in rows if r.allocation_effect_pct is not None]
    relation = "贏過" if alloc_total >= 0 else "落後"
    if not effects:
        return f"產業配置對相對表現的影響約 {alloc_total:+.2f} 個百分點。"
    best = max(effects, key=lambda r: r.allocation_effect_pct or 0)
    worst = min(effects, key=lambda r: r.allocation_effect_pct or 0)
    return (
        f"光看產業配置差異，今天{relation}{benchmark_name} {alloc_total:+.2f} 個百分點。"
        f"最有利的是 {best.sector}（配置差異 {best.weight_diff_pct:+.1f}%、"
        f"當日 {best.sector_return_pct:+.2f}%）；最不利的是 {worst.sector}。"
    )


def _read_tabular_text(text: str) -> tuple[list[dict[str, str]], list[str]]:
    lines = [line for line in text.replace("\ufeff", "").splitlines() if line.strip()]
    if not lines:
        return [], []

    delimiter = _guess_delimiter(lines)
    reader = csv.reader(io.StringIO("\n".join(lines)), delimiter=delimiter)
    raw_rows = [row for row in reader if any(cell.strip() for cell in row)]
    if not raw_rows:
        return [], []

    header_index, field_map = _find_header(raw_rows)
    if header_index is None:
        return [], []

    headers = [cell.strip() for cell in raw_rows[header_index]]
    parsed_rows: list[dict[str, str]] = []
    for raw in raw_rows[header_index + 1 :]:
        normalized: dict[str, str] = {}
        for source_idx, target in field_map.items():
            if source_idx < len(raw):
                normalized[target] = raw[source_idx].strip()
        if normalized:
            parsed_rows.append(normalized)
    return parsed_rows, headers


def _read_workbook_holdings(
    content: bytes,
) -> tuple[list[dict[str, str]], list[str], dict[str, str], list[str]]:
    workbook = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    warnings: list[str] = []
    fallback: tuple[list[dict[str, str]], list[str], dict[str, str], list[str]] | None = None

    for sheet in workbook.worksheets:
        raw_rows = [
            ["" if cell is None else str(cell).strip() for cell in row]
            for row in sheet.iter_rows(values_only=True)
            if any(cell is not None and str(cell).strip() for cell in row)
        ]
        if not raw_rows:
            continue

        header_index, field_map = _find_header(raw_rows)
        if header_index is None:
            continue

        headers = [cell.strip() for cell in raw_rows[header_index]]
        parsed_rows = _rows_after_header(raw_rows, header_index, field_map)
        metadata = _workbook_metadata(sheet.title, raw_rows)
        result = (parsed_rows, headers, metadata, warnings)
        if "股票" in sheet.title:
            return result
        if fallback is None:
            fallback = result

    if fallback is not None:
        warnings.append("沒有找到名稱包含「股票」的工作表，已改用第一個可解析工作表。")
        return fallback

    return [], [], {}, ["找不到包含股票代碼、股票名稱、權重欄位的工作表。"]


def _rows_after_header(
    raw_rows: list[list[str]],
    header_index: int,
    field_map: dict[int, str],
) -> list[dict[str, str]]:
    parsed_rows: list[dict[str, str]] = []
    for raw in raw_rows[header_index + 1 :]:
        normalized: dict[str, str] = {}
        for source_idx, target in field_map.items():
            if source_idx < len(raw):
                normalized[target] = raw[source_idx].strip()
        if normalized:
            parsed_rows.append(normalized)
    return parsed_rows


def _workbook_metadata(sheet_title: str, raw_rows: list[list[str]]) -> dict[str, str]:
    metadata: dict[str, str] = {}
    for row in raw_rows[:8]:
        first = row[0].strip() if row else ""
        if not first:
            continue
        if not metadata.get("as_of"):
            match = re.search(r"\((\d{4}-\d{2}-\d{2})\)", first)
            if match:
                metadata["as_of"] = match.group(1)
        if (
            not metadata.get("fund_name")
            and "基金" in first
            and not first.startswith(("(", "基金資產"))
        ):
            metadata["fund_name"] = first

    if not metadata.get("as_of"):
        match = re.search(r"\((\d{4}-\d{2}-\d{2})\)", sheet_title)
        if match:
            metadata["as_of"] = match.group(1)
    return metadata


def _guess_delimiter(lines: list[str]) -> str:
    candidates = [",", "\t", ";", "|"]
    sample = "\n".join(lines[:8])
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters="".join(candidates))
        return dialect.delimiter
    except csv.Error:
        return max(
            candidates,
            key=lambda delimiter: sum(line.count(delimiter) for line in lines[:8]),
        )


def _find_header(rows: list[list[str]]) -> tuple[int | None, dict[int, str]]:
    best_index: int | None = None
    best_map: dict[int, str] = {}
    best_score = 0
    for index, row in enumerate(rows[:20]):
        field_map: dict[int, str] = {}
        for cell_index, cell in enumerate(row):
            target = _field_for_header(cell)
            if target:
                field_map[cell_index] = target
        score = len(set(field_map.values()))
        if score > best_score and {"name", "weight_pct"}.issubset(set(field_map.values())):
            best_index = index
            best_map = field_map
            best_score = score
    return best_index, best_map


def _field_for_header(value: str) -> str | None:
    normalized = _normalize_header(value)
    for target, aliases in COLUMN_ALIASES.items():
        if normalized in aliases:
            return target
    return None


def _normalize_header(value: str) -> str:
    cleaned = value.strip().lower()
    cleaned = cleaned.replace("%", " percent ")
    cleaned = re.sub(r"[\s/().\-]+", "_", cleaned)
    cleaned = re.sub(r"_+", "_", cleaned).strip("_")
    return cleaned


def _clean_text(value: str | None) -> str:
    return (value or "").strip().strip('"').strip()


def _parse_percent(value: str | None) -> float | None:
    cleaned = _clean_text(value)
    if not cleaned or cleaned in {"-", "--", "N/A", "n/a"}:
        return None
    cleaned = cleaned.replace("%", "").replace(",", "").replace("％", "").strip()
    if cleaned.startswith("(") and cleaned.endswith(")"):
        cleaned = f"-{cleaned[1:-1]}"
    try:
        return round(float(cleaned), 6)
    except ValueError:
        return None


def _parse_iso_date(value: str) -> date | None:
    try:
        return date.fromisoformat(value)
    except ValueError:
        return None


def _normalize_twse_symbol(value: str) -> str:
    match = re.search(r"\b[0-9]{4,6}\b", value)
    return match.group(0) if match else ""


def _normalize_twse_fund_symbol(value: str) -> str:
    # Active ETF codes can carry a trailing letter (e.g. 00982A); keep it. No \b
    # after the digits because a digit→letter transition is not a word boundary.
    match = re.search(r"[0-9]{4,6}[A-Za-z]?", value.strip())
    return match.group(0).upper() if match else ""


def _infer_symbol(name: str) -> str:
    match = re.search(r"\b[0-9]{4,6}\b", name)
    if match:
        return match.group(0)
    return ""


def _row_from_holding(holding: HoldingInput) -> AttributionRow:
    if holding.return_pct is None:
        return AttributionRow(
            symbol=holding.symbol,
            name=holding.name,
            weight_pct=round(holding.weight_pct, 4),
            return_pct=None,
            contribution_pct=None,
            direction="missing",
        )

    contribution = holding.weight_pct * holding.return_pct / 100
    if contribution > 0:
        direction = "positive"
    elif contribution < 0:
        direction = "negative"
    else:
        direction = "flat"
    return AttributionRow(
        symbol=holding.symbol,
        name=holding.name,
        weight_pct=round(holding.weight_pct, 4),
        return_pct=round(holding.return_pct, 4),
        contribution_pct=round(contribution, 4),
        direction=direction,
    )


def _summary_zh(
    *,
    fund_name: str,
    benchmark_name: str,
    fund_return: float,
    benchmark_return: float,
    active_return: float,
    contributors: list[AttributionRow],
    drags: list[AttributionRow],
) -> str:
    relation = "贏過" if active_return >= 0 else "落後"
    leader = contributors[0].name if contributors else "主要持股"
    laggard = drags[0].name if drags else "部分持股或現金部位"
    return (
        f"{fund_name} 當日報酬 {fund_return:+.2f}%，{benchmark_name} {benchmark_return:+.2f}%，"
        f"相對表現 {active_return:+.2f} 個百分點，今天{relation}基準。"
        f"主要正貢獻來自 {leader}；主要拖累來自 {laggard}。"
    )
