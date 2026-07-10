"""XLSX source workbook (plan §3: raw-data, source, and ledger tabs).

Security (plan §9): every string cell passes through _safe_cell — values starting
with = + - @ or tab/CR are prefixed with ' so spreadsheet apps treat them as text,
never formulas. LLM- or filing-influenced text cannot become =WEBSERVICE() exfil.
"""

import io

from app.exports.bundle import ExportBundle, strip_claim_refs

_FORMULA_PREFIXES = ("=", "+", "-", "@", "\t", "\r")


def _safe_cell(value):
    if isinstance(value, str) and value.startswith(_FORMULA_PREFIXES):
        return "'" + value
    return value


def build_xlsx(bundle: ExportBundle, time_series: list[dict] | None = None) -> bytes:
    from openpyxl import Workbook  # noqa: PLC0415
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    wb = Workbook()
    head_font = Font(bold=True, size=10, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="00477B")

    def style_header(ws, ncols: int) -> None:
        for col in range(1, ncols + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = head_font
            cell.fill = head_fill
        ws.freeze_panes = "A2"

    def autosize(ws, widths: list[int]) -> None:
        for i, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = w

    # --- Brief tab ---
    ws = wb.active
    ws.title = "Brief"
    ws.append(["Watermark", _safe_cell(bundle.watermark)])
    ws.append(["Watchlist", _safe_cell(bundle.watchlist)])
    ws.append(["Generated", bundle.generated_at.strftime("%Y-%m-%d %H:%M UTC")])
    ws.append(["Model", _safe_cell(bundle.model)])
    ws.append(["Prompt version", _safe_cell(bundle.prompt_version)])
    ws.append(["Status", _safe_cell(bundle.status)])
    if bundle.approval_line:
        ws.append(["Approval", _safe_cell(bundle.approval_line)])
    ws.append(["AI generated (Art. 50)", "true"])
    ws.append([])
    ws.append(["Section", "Content", "Review action"])
    header_row = ws.max_row
    for section in bundle.sections:
        ws.append(
            [
                _safe_cell(section.title),
                _safe_cell(strip_claim_refs(section.content)),
                _safe_cell(section.action or ""),
            ]
        )
    for col in range(1, 4):
        ws.cell(row=header_row, column=col).font = Font(bold=True)
    autosize(ws, [24, 110, 14])

    # --- Evidence ledger tab ---
    ws = wb.create_sheet("Evidence Ledger")
    ws.append(["ID", "Claim", "Type", "Confidence", "Status", "Sources", "Reason", "Span IDs"])
    style_header(ws, 8)
    for claim in bundle.supported_claims + bundle.review_claims:
        ws.append(
            [
                _safe_cell(claim.cid),
                _safe_cell(claim.text),
                _safe_cell(claim.claim_type),
                _safe_cell(claim.confidence),
                _safe_cell(claim.support_status),
                _safe_cell("; ".join(claim.sources)),
                _safe_cell(claim.reason),
                _safe_cell("; ".join(m.get("span_id", "") for m in claim.meta)),
            ]
        )
    autosize(ws, [8, 70, 16, 12, 12, 50, 36, 40])

    # --- Sources tab ---
    ws = wb.create_sheet("Sources")
    ws.append(["Claim", "Doc type", "Accession", "Section", "Span", "Source URL", "Retrieved", "SHA-256"])
    style_header(ws, 8)
    for claim in bundle.supported_claims:
        for m in claim.meta:
            span = m.get("span") or [None, None]
            ws.append(
                [
                    _safe_cell(claim.cid),
                    _safe_cell(m.get("doc_type") or ""),
                    _safe_cell(m.get("accession") or ""),
                    _safe_cell(m.get("section") or ""),
                    _safe_cell(f"{span[0]}–{span[1]}" if span[0] is not None else ""),
                    _safe_cell(m.get("source_url") or ""),
                    _safe_cell(m.get("retrieved_at") or ""),
                    _safe_cell(m.get("checksum_sha256") or ""),
                ]
            )
    autosize(ws, [8, 10, 24, 12, 14, 60, 22, 40])

    # --- Macro data tab (raw observations, vintage-stamped) ---
    if time_series:
        ws = wb.create_sheet("Macro Data")
        ws.append(["Series", "Date", "Value", "Units", "Frequency", "Vintage"])
        style_header(ws, 6)
        for series in time_series:
            for date, value in sorted((series.get("observations") or {}).items()):
                try:
                    numeric = float(value)
                except (TypeError, ValueError):
                    numeric = None
                ws.append(
                    [
                        _safe_cell(series.get("series_id", "")),
                        _safe_cell(date),
                        numeric if numeric is not None else _safe_cell(str(value)),
                        _safe_cell(series.get("units", "")),
                        _safe_cell(series.get("frequency", "")),
                        _safe_cell(series.get("vintage", "")),
                    ]
                )
        autosize(ws, [14, 12, 12, 28, 12, 22])

    # --- Disclosures tab ---
    ws = wb.create_sheet("Disclosures")
    ws.append(["Disclosure"])
    style_header(ws, 1)
    ws.append([_safe_cell(bundle.disclaimer)])
    ws.append([f"ai_generated=true; brief_id={bundle.brief_id}; format=cited-market-brief-agent.xlsx/v1"])
    autosize(ws, [140])

    wb.properties.description = f"ai_generated=true; brief_id={bundle.brief_id}; model={bundle.model}"

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
